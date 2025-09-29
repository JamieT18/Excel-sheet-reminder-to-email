import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook

EXCEL_PATH = 'license_dates.xlsx'  # Change to your actual file location

def safe_read_excel(path):
    try:
        df = pd.read_excel(path, header=0)
        return df
    except Exception as e:
        messagebox.showerror("File Error", f"Could not read Excel file:\n{e}")
        return None

def update_due_dates(df):
    today = datetime.today()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    updated = False

    for idx, row in df.iterrows():
        for col_idx, cell in enumerate(row):
            if isinstance(cell, str) and cell.startswith('Due'):
                parts = cell.split()
                if len(parts) == 2:
                    date_str = parts[1]
                    try:
                        due_date = datetime.strptime(date_str, '%m.%d.%Y')
                        if (due_date.month, due_date.day) == (today.month, today.day):
                            # Update to next year
                            next_year_date = due_date.replace(year=due_date.year + 1)
                            new_cell = f'Due {next_year_date.strftime("%m.%d.%Y")}'
                            ws.cell(row=idx + 2, column=col_idx + 1, value=new_cell)
                            updated = True
                    except Exception:
                        continue
    if updated:
        wb.save(EXCEL_PATH)

def get_dates(df):
    today = datetime.today()
    date_list = []
    for idx, row in df.iterrows():
        location = row[0]
        due_date_str = ""
        expires_date_str = ""
        highlight_due = False
        highlight_expires = False
        for cell in row[1:]:
            if isinstance(cell, str) and cell.startswith('Due'):
                parts = cell.split()
                if len(parts) == 2:
                    date_str = parts[1]
                    try:
                        due_date = datetime.strptime(date_str, '%m.%d.%Y')
                        due_date_str = date_str
                        highlight_due = (due_date.month, due_date.day) == (today.month, today.day)
                    except Exception:
                        continue
            elif isinstance(cell, str) and cell.startswith('Expires'):
                parts = cell.split()
                if len(parts) == 2:
                    date_str = parts[1]
                    try:
                        expires_date = datetime.strptime(date_str, '%m.%d.%Y')
                        expires_date_str = date_str
                        highlight_expires = (expires_date.month, expires_date.day) == (today.month, today.day)
                    except Exception:
                        continue
        date_list.append((location, due_date_str, expires_date_str, highlight_due, highlight_expires))
    return date_list

def refresh_ui(tree, status_bar):
    df = safe_read_excel(EXCEL_PATH)
    if df is None:
        return
    update_due_dates(df)
    for item in tree.get_children():
        tree.delete(item)
    dates = get_dates(df)
    for location, due_date, expires_date, highlight_due, highlight_expires in dates:
        tags = ()
        if highlight_due:
            tags += ('highlight_due',)
        if highlight_expires:
            tags += ('highlight_expires',)
        tree.insert('', 'end', values=(location, due_date, expires_date), tags=tags)
    tree.tag_configure('highlight_due', background='yellow')
    tree.tag_configure('highlight_expires', background='tomato')
    status_bar.config(text=f"Last refresh: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Locations: {len(dates)}")

def sort_column(tree, col, reverse):
    l = [(tree.set(k, col), k) for k in tree.get_children('')]
    try:
        l.sort(key=lambda t: datetime.strptime(t[0], "%m.%d.%Y"), reverse=reverse)
    except:
        l.sort(reverse=reverse)
    for index, (val, k) in enumerate(l):
        tree.move(k, '', index)
    tree.heading(col, command=lambda: sort_column(tree, col, not reverse))

def create_ui():
    root = tk.Tk()
    root.title("License Manager")

    frame = ttk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True)

    columns = ('Location', 'Due Date', 'Expires Date')
    tree = ttk.Treeview(frame, columns=columns, show='headings', height=30)
    for col in columns:
        tree.heading(col, text=col, command=lambda _col=col: sort_column(tree, _col, False))
        tree.column(col, width=180, anchor='center')
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    vsb.pack(side=tk.RIGHT, fill=tk.Y)

    search_var = tk.StringVar()
    search_entry = ttk.Entry(root, textvariable=search_var, width=30)
    search_entry.pack(side=tk.LEFT, padx=5, pady=5)
    search_entry.insert(0, "Search location...")

    def on_search(*args):
        filter_text = search_var.get().strip().lower()
        df = safe_read_excel(EXCEL_PATH)
        if df is None:
            return
        update_due_dates(df)
        for item in tree.get_children():
            tree.delete(item)
        dates = [d for d in get_dates(df) if filter_text in d[0].lower()]
        for location, due_date, expires_date, highlight_due, highlight_expires in dates:
            tags = ()
            if highlight_due:
                tags += ('highlight_due',)
            if highlight_expires:
                tags += ('highlight_expires',)
            tree.insert('', 'end', values=(location, due_date, expires_date), tags=tags)
        tree.tag_configure('highlight_due', background='yellow')
        tree.tag_configure('highlight_expires', background='tomato')
        status_bar.config(text=f"Filtered: {len(dates)} locations")

    search_var.trace_add("write", on_search)

    refresh_btn = ttk.Button(root, text="Refresh", command=lambda: refresh_ui(tree, status_bar))
    refresh_btn.pack(side=tk.LEFT, padx=5, pady=5)

    status_bar = ttk.Label(root, text="", anchor='w')
    status_bar.pack(fill=tk.X, padx=5)

    refresh_ui(tree, status_bar)
    root.mainloop()

if __name__ == "__main__":
    create_ui()
